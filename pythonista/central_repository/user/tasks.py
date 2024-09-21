from __future__ import absolute_import, unicode_literals
from datetime import datetime
import json, os, string
from openpyxl import load_workbook
from random import choice

from django.db import transaction
from django.db.models import Q
from django.core.cache import cache
from django.core.mail import EmailMultiAlternatives, get_connection
from django.core.mail import send_mail, get_connection
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings

from config.celery import app
from hostel.models import Block, Hostel, Level, Room, RoomType
from user.models import User, Student


@app.task(name="send_password_reset_mail")
def send_password_reset_mail(name, url, email):
    idx = cache.get("emailIdIndex", 0)
    cache.set("emailIdIndex", (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)

    connection = get_connection(
        username=settings.EMAIL_HOST_USERS[idx],
        password=settings.EMAIL_HOST_PASSWORDS[idx],
        fail_silently=False,
    )
    connection.open()

    subject = f"Password Reset Link of Online Hostel Management System for {name}"

    context = {
        "name": name,
        "url": url,
    }
    html_message = render_to_string("user/password_reset.html", context)
    msg = strip_tags(html_message)
    email_message = EmailMultiAlternatives(
        subject,
        msg,
        settings.EMAIL_HOST_USERS[idx],
        [email],
        reply_to=["queries_studentaffairs@thapar.edu"],
        connection=connection,
    )
    email_message.attach_alternative(html_message, "text/html")
    email_message.send()
    # send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (email, ), html_message=html_message, connection=connection, fail_silently=False)
    connection.close()

    return f"\n password reset mail sent to {email}\n"


@app.task(name="import_students")
def import_students(filename):
    start_time = datetime.now()

    path = os.path.join(settings.BASE_DIR, "imported_data", "student", filename)
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    log_file_path = os.path.join(settings.LOGS_ROOT, "import_students.log")
    with open(log_file_path, "a") as f:
        f.write(f"\nProcessing file {filename} at {start_time}...\n")

    # Check if all fields are present in given data
    # for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=11):
    #     for i in range(len(row)):
    #         if row[i].value != studentfileformat[i]:
    #             with open(log_file_path, "a") as f:
    #                 f.write(
    #                     f"Invalid File Format. Found {row[i].value}, but required was {studentfileformat[i]}\n"
    #                 )
    #                 f.write(f"Correct File Format is {str(studentfileformat)}\n")
    #                 f.write("Aborting...\n\n")
    #                 f.write("\n")
    #             raise Exception(
    #                 f"Found {row[i]}, but required was {studentfileformat[i]}\n"
    #             )

    successCnt = 0
    failureCnt = 0
    cnt = 1
    errors = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=11):
        try:
            with transaction.atomic():
                student = None

                # Extract data into variables and data preprocessing
                rollno = row[0].value
                if rollno is not None:
                    student = Student.objects.filter(roll_number=rollno.strip()).first()

                email = row[1].value
                if email is not None:
                    email = email.strip()
                    if student is not None:
                        user = student.user
                        user.email = email
                        user.save()
                    else:
                        user = User.objects.filter(email=email).first()
                        if user is None:
                            user = User(email=email)
                            password = "".join(
                                choice(
                                    string.ascii_letters
                                    + string.digits
                                    + "!@#$%^&*()[]{};:?.,<>_"
                                )
                                for _ in range(8)
                            )
                            user.set_password(password)
                            user.save()
                        try:
                            if user.student is not None:
                                student = user.student
                        except:
                            pass

                # room
                room = None
                room_name = row[4].value
                if room_name is not None:
                    room_name = room_name.strip()
                    if room_name != "":
                        # check room data if given

                        # Room type and Hostel
                        roomtype_name = row[5].value
                        hostel_name = row[6].value
                        if hostel_name is not None:
                            hostel_name = hostel_name.strip()
                            hostel = Hostel.objects.filter(
                                name__iexact=hostel_name
                            ).first()
                            if hostel is None:
                                raise Exception(
                                    f"Current Hostel {hostel_name} not found!"
                                )
                            roomtype = RoomType.objects.filter(
                                Q(name__iexact=roomtype_name) & Q(hostel=hostel)
                            ).first()
                            if roomtype is None:
                                raise Exception(
                                    f"Current Room Type {roomtype_name} not found for Hostel {hostel_name}!"
                                )
                        else:
                            raise Exception("Hostel is NULL!")

                        room_repr_arr = room_name.split("-")
                        if len(room_repr_arr) != 2:
                            raise Exception(
                                f"Invalid format for room name - {room_name}. It should be Block-RoomName"
                            )

                        block_name = room_repr_arr[0].strip()
                        room_name_short = room_repr_arr[1].strip()

                        block = Block.objects.filter(
                            name=block_name, hostel=hostel
                        ).first()
                        if block is None:
                            raise Exception(
                                f"No block found with name {block_name} in hostel {hostel_name}"
                            )

                        level_name = (
                            "G"
                            if len(room_name_short) < 3 or room_name_short[0] == "0"
                            else room_name_short[0]
                        )
                        level = Level.objects.filter(
                            name=level_name, block=block
                        ).first()
                        if level is None:
                            raise Exception(
                                f"No level found with name {level_name} in block {block_name} of hostel {hostel_name}"
                            )

                        room = Room.objects.filter(
                            name=room_name, level=level, room_type=roomtype
                        ).first()
                        if room is None:
                            raise Exception(
                                f"No room found with name {room_name} in level {level_name} of block {block_name} of hostel {hostel_name}"
                            )

                name = row[2].value
                if name is not None:
                    name = name.strip()

                gender = row[3].value
                if gender is not None:
                    gender = gender.strip()

                if student is None:
                    # create new student
                    student = Student.objects.create(
                        name=name,
                        roll_number=rollno,
                        gender=gender,
                        user=user,
                        room=room,
                    )
                else:
                    # update details of existing student
                    if name is not None:
                        student.name = name
                    if rollno is not None:
                        student.roll_number = rollno
                    if gender is not None:
                        student.gender = gender
                    if room is not None:
                        student.room = room
                    student.save()

                successCnt += 1

        except Exception as e:
            # log error
            errors.append(f"Creation of item {cnt} unsuccessful! Error: {str(e)}")
            failureCnt += 1

        cnt += 1

    res = {"successful": successCnt, "unsuccessful": failureCnt, "errors": errors}

    end_time = datetime.now()
    time_diff = (end_time - start_time).total_seconds() / 60

    with open(log_file_path, "a") as f:
        f.write(f"Execution completed in {time_diff} minutes.\n")
        f.write(json.dumps(res, indent=2))
        f.write("\n")

    errors_str = "\n- ".join(errors)

    subject = f"Execution results ready for the task Import Students"
    msg = f"""File: {filename}

Task Start Time: {start_time}
Task End Time: {end_time}
Time Elapsed: {time_diff}

Number of entries successfully executed: {successCnt}
Number entries failed: {failureCnt}

Errors
- {'No Errors' if len(errors)==0 else errors_str}
"""

    idx = cache.get("emailIdIndex", 0)
    cache.set("emailIdIndex", (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)

    connection = get_connection(
        username=settings.EMAIL_HOST_USERS[idx],
        password=settings.EMAIL_HOST_PASSWORDS[idx],
        fail_silently=False,
    )
    connection.open()

    send_mail(
        subject,
        msg,
        settings.EMAIL_HOST_USERS[idx],
        (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL),
        connection=connection,
        fail_silently=False,
    )

    connection.close()

    return res
