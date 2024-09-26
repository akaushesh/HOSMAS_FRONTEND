from datetime import datetime
from openpyxl import load_workbook
from random import choice
import string, os

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.status import HTTP_200_OK, HTTP_202_ACCEPTED, HTTP_400_BAD_REQUEST
from rest_framework.response import Response

from django.core.files.storage import default_storage
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist

from . import services as user_services
from .models import Student, User
from .serializers import StudentProfileSerializer
from .tasks import import_students, send_password_reset_mail

from common.permissions import IsSupervisor
import common.services as common_services


class ProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(
            user_services.get_user_detailed_profile(request.user.id), status=HTTP_200_OK
        )


class getStudentProfile(APIView):
    permission_classes = [IsAuthenticated, IsSupervisor]

    def get(self, request, student_id):
        try:
            student_instance = common_services.get_object(
                Student.objects, id=student_id
            )
        except ObjectDoesNotExist:
            return Response(
                {"detail": f"No student found with id {student_id}"},
                status=HTTP_400_BAD_REQUEST,
            )
        serializer = StudentProfileSerializer(student_instance)
        return Response(serializer.data, status=HTTP_200_OK)


class EmailResetPasswordView(APIView):
    def post(self, request):
        user = common_services.filter_objects(
            User.objects, email=request.data.get("email")
        ).first()
        if user is None:
            return Response(status=HTTP_400_BAD_REQUEST)

        if user.password_reset_slug is None or user.password_reset_slug == "":
            while True:
                slug = "".join(
                    choice(string.ascii_letters + string.digits + "-")
                    for _ in range(135)
                )
                if not User.objects.filter(password_reset_slug=slug).exists():
                    break
            user.password_reset_slug = slug
            user.save()

        try:
            name = user.student.name
        except:
            try:
                name = user.supervisor.name
            except:
                name = "Admin"

        url = (
            f"{settings.STUDENT_PORTAL_URL}/auth/reset-password/"
            + user.password_reset_slug
        )
        send_password_reset_mail.delay(name, url, user.email)

        return Response(status=HTTP_200_OK)


class ResetPasswordView(APIView):
    def post(self, request):
        user = common_services.filter_objects(
            User.objects, password_reset_slug=request.data.get("slug")
        ).first()
        password = request.data.get("password")

        if user is None:
            return Response({"detail": "Invalid Slug!"}, status=HTTP_400_BAD_REQUEST)

        if password is None or len(password) < 8:
            return Response(
                {"detail": "Password should be atleast 8 characters long!"},
                status=HTTP_400_BAD_REQUEST,
            )

        user.set_password(password)
        user.password_reset_slug = None
        user.save()

        return Response(status=HTTP_200_OK)


class ImportStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    studentfileformat = (
        "roll number",
        "email",
        "name",
        "gender",
        "room",
        "room type",
        "hostel",
    )

    def post(self, request):
        file = request.data.get("file")

        if file is None:
            return Response({"detail": "File Not Found!"}, status=HTTP_400_BAD_REQUEST)

        filename = f"{datetime.now().strftime('%Y%m%d_%H%M')}_{file.name}"

        if filename.split(".")[-1] != "xlsx":
            return Response(
                {"detail": "Only xlsx file format supported!"},
                status=HTTP_400_BAD_REQUEST,
            )

        storage = default_storage
        storage.location = os.path.join(settings.BASE_DIR, "imported_data", "student")
        filename = storage.save(filename, file)

        path = os.path.join(settings.BASE_DIR, "imported_data", "student", filename)
        wb = load_workbook(path)
        ws = wb.active

        # Check if all fields are present in given data
        for row in ws.iter_rows(min_row=1, max_row=1):
            if len(row) != len(self.studentfileformat):
                return Response(
                    {
                        "detail": f"Invalid File Format. Correct File Format is {str(self.studentfileformat)}. File's header length is {len(row)}"
                    },
                    status=HTTP_400_BAD_REQUEST,
                )
            for i in range(len(row)):
                if row[i].value.strip().lower() != self.studentfileformat[i]:
                    return Response(
                        {
                            "detail": f"Invalid File Format. Found {row[i].value}, but required was {self.studentfileformat[i]}, Correct File Format is {str(self.studentfileformat)}"
                        },
                        status=HTTP_400_BAD_REQUEST,
                    )

        import_students.delay(filename)

        return Response(status=HTTP_202_ACCEPTED)
