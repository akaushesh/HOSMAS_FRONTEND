from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.conf import settings
from django.db import transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from student.permissions import IsStudent
from student.tasks import *

from .permissions import IsAdmin

from preference.models import Hostel, RoomType, RoomTypeChoice, Room
from student.models import Batch, Section, Student, Group
from .models import AllotmentStatus, AcademicSession, Faq

from .serializers import HostelSerializer, HostelSingleSerializer, RoomTypeSerializer, RoomTypeChoiceSerializer, RoomTypeOptionSerializer, BatchSerializer, BatchUninitializedSerializer, SectionSerializer, ProfileSerializer, AllotmentStatusSerializer, SectionRoomTypeSerializer, AcademicSessionSerializer, FAQSerializer, DefaulterSerializer
from .serializers import *
from student.serializers import StudentSerializer, GroupSerializer, StudentProfileSerializer, StudentSerializer

from .tasks import send_reminder_mail, import_students, import_defaulters, allot, studentfileformat

from datetime import datetime
import csv, os, json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from collections import OrderedDict

from preference.models import Preference

# Create your views here.


class CreateObjectView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def post(self, request, model):
            if model=='hostel':
                  serializer = HostelSingleSerializer(data=request.data)
            elif model=='roomtype':
                  serializer = RoomTypeSerializer(data=request.data)
            elif model  == 'room':
                  serializer = RoomSerializer(data=request.data)
            elif model=='choice':
                  serializer = RoomTypeChoiceSerializer(data=request.data)
            elif model=='batch':
                  serializer = BatchSerializer(data = request.data)
            elif model=='section':
                  serializer = SectionSerializer(data = request.data)
            elif model=='defaulter':
                  student = Student.objects.filter(rollno=request.data.get('student')).first()
                  if student is None:
                        return Response(status=status.HTTP_400_BAD_REQUEST)
                  serializer = DefaulterSerializer(data = {'student': student.id})
            elif model=='student':
                  serializer = StudentProfileSerializer(data=request.data, context={'is_admin': True})
            else:
                  return Response(status=status.HTTP_404_NOT_FOUND)
            
            if not serializer.is_valid():
                  return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()

            return Response(serializer.data, status=status.HTTP_201_CREATED)


class GetMultipleObjectsView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def get(self, request, model):
            if model=='hostel':
                  queryset = Hostel.objects.all()
                  serializer = HostelSerializer(queryset, many=True)
            elif model=='section':
                  queryset = Section.objects.select_related('batch').all()
                  serializer = SectionSerializer(queryset, many=True)
            elif model=='uninitialized-batch':
                  queryset = Batch.objects.annotate(
                        sections_count = Count('sections')
                  ).filter(sections_count__lt=2)
                  serializer = BatchUninitializedSerializer(queryset, many=True)
            elif model=='batch':
                  queryset = Batch.objects.all()
                  serializer = BatchSerializer(queryset, many=True)
            elif model=='roomtype':
                  queryset = RoomType.objects.select_related('hostel').all()
                  serializer = RoomTypeOptionSerializer(queryset, many=True)
            elif model=='choice':
                  queryset = Section.objects.filter(id=request.GET.get('section')).select_related('batch').prefetch_related('choices').first()
                  if queryset is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = SectionRoomTypeSerializer(queryset)
            elif model=='allotment-status':
                  queryset = AllotmentStatus.objects.prefetch_related('sections', 'pending_groups__members__preview_room__hostel', 'pending_groups__leader__preview_room__hostel').first()
                  if queryset is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = AllotmentStatusSerializer(queryset, many=True)
            else:
                  return Response(status=status.HTTP_404_NOT_FOUND)
            return Response(serializer.data, status=status.HTTP_200_OK)


class GetObjectView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]
      
      def get(self, request, model, id):
            if model=='hostel':
                  instance = Hostel.objects.filter(id=id).prefetch_related('room_types').first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = HostelSingleSerializer(instance)
            
            elif model=='roomtype':
                  instance = RoomType.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = RoomTypeSerializer(instance)
            
            elif model=='allotment-status':
                  instance = AllotmentStatus.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = AllotmentStatusSerializer(instance)
            
            elif model=='academic-session':
                  instance = AcademicSession.objects.first()
                  if instance is None:
                        instance = AcademicSession(name='')
                        instance.save()
                  serializer = AcademicSessionSerializer(instance)
            
            elif model=='student':
                  instance = Student.objects.filter(rollno=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = StudentProfileSerializer(instance, context={'is_admin': True})
            
            elif model=='search-student':
                  # View to search the student to add it into a group
                  instance = Student.objects.filter(rollno=id).first()
                  
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  
                  try:
                        _ = instance.defaulter
                        return Response({"detail": "This student is suspended from Hostel Allocation Process"}, status=status.HTTP_403_FORBIDDEN)
                  except ObjectDoesNotExist:
                        pass
                  
                  try:
                        gp = instance.leader_of_group
                        if gp.members.count()>0:
                              return Response({'detail': 'This student is already a leader of another group!'}, status=status.HTTP_403_FORBIDDEN)
                  except ObjectDoesNotExist:
                        if instance.group is not None:
                              return Response({'detail': 'This student is already a part of another group!'}, status=status.HTTP_409_CONFLICT)
                  
                  serializer = StudentSerializer(instance)
            
            elif model=='batch':
                  instance = Batch.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = BatchSerializer(instance)
            
            else:
                  return Response(status.HTTP_404_NOT_FOUND)
            
            return Response(serializer.data, status=status.HTTP_200_OK)


class UpdateObjectView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def put(self, request, model, id):
            if model=='hostel':
                  instance = Hostel.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = HostelSingleSerializer(instance, request.data)
            elif model=='roomtype':
                  instance = RoomType.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = RoomTypeSerializer(instance, request.data)
            elif model=='choice':
                  instance = RoomTypeChoice.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = RoomTypeChoiceSerializer(instance, request.data, partial=True)
            elif model=='section':
                  instance = Section.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  fee_submission_deadline = request.data.get('fee_submission_deadline', '')
                  reporting_information = request.data.get('reporting_information', '')
                  context = {
                        'fee_submission_deadline': fee_submission_deadline,
                        'reporting_information': reporting_information
                  }
                  serializer = SectionSerializer(instance, request.data, context=context, partial=True)
            elif model=='batch':
                  instance = Batch.objects.filter(id=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = BatchSerializer(instance, request.data)
            elif model=='allotment-status':
                  instance = AllotmentStatus.objects.first()
                  if instance is None:
                        instance = AllotmentStatus()
                        instance.save()
                  serializer = AllotmentStatusSerializer(instance, request.data)
            elif model=='academic-session':
                  instance = AcademicSession.objects.first()
                  if instance is None:
                        instance = AcademicSession(name='')
                        instance.save()
                  serializer = AcademicSessionSerializer(instance, request.data)
            elif model=='student':
                  instance = Student.objects.filter(rollno=id).first()
                  if instance is None:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  serializer = StudentProfileSerializer(instance, request.data, partial=True, context={'is_admin': True})
            else:
                  return Response(status=status.HTTP_404_NOT_FOUND)

            if not serializer.is_valid():
                  return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()

            return Response(serializer.data, status=status.HTTP_200_OK)


class DeleteObjectView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def delete(self, request, model):
            id = request.data.get('id')
            if id is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)

            if model=='hostel':
                  instance = Hostel.objects.filter(id=id).first()
            elif model=='roomtype':
                  instance = RoomType.objects.filter(id=id).first()
            elif model=='choice':
                  instance = RoomTypeChoice.objects.filter(id=id).first()
            elif model=='section':
                  instance = Section.objects.filter(id=id).first()
            elif model=='batch':
                  instance = Batch.objects.filter(id=id).first()
            elif model=='section':
                  instance = Section.objects.filter(id=id).first()
            elif model=='defaulter':
                  instance = Defaulter.objects.filter(id=id).first()
            elif model=='student':
                  instance = Student.objects.filter(rollno=id).first()
            else:
                  return Response(status=status.HTTP_404_NOT_FOUND)

            if instance is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)

            instance.delete()
            return Response(status=status.HTTP_200_OK)


class DeleteObjectsView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def delete(self, request, model):
            ids = request.data.get('ids')
            for id in ids:
                  if model=='defaulter':
                        instance = Defaulter.objects.filter(id=id).first()
                  elif model=='student':
                        instance = Student.objects.filter(rollno=id).first()
                  else:
                        return Response(status=status.HTTP_404_NOT_FOUND)
                  if instance is not None:
                        instance.delete()
            return Response(status=status.HTTP_200_OK)


class getStudents(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]
      
      def get(self, request):
            q = request.GET.get('q')
            batch_id = request.GET.get('batch')

            students_per_page = request.GET.get('students_per_page')
            
            if batch_id=='all':
                  if q is not None and q.strip()!='':
                        students_list = Student.objects.filter(Q(rollno__startswith = q) | Q(name__icontains = q) | Q(user__email__icontains = q) ).distinct().order_by('id').all()
                        total_cnt = Student.objects.filter(Q(rollno__startswith = q) | Q(name__icontains = q) | Q(user__email__icontains = q) ).distinct().order_by('id').count()
                  else:
                        students_list = Student.objects.order_by('id').all()
                        total_cnt = Student.objects.order_by('id').count()
            else:
                  batch = Batch.objects.filter(id = batch_id).first()
                  if batch is None:
                        return Response({'error':'Batch does not exist'}, status=status.HTTP_404_NOT_FOUND)
                  if q is not None and q.strip()!='':
                        students_list = Student.objects.filter(Q(batch = batch), Q(rollno__startswith = q) | Q(name__icontains = q) | Q(user__email__icontains = q) ).distinct().order_by('id').all()
                        total_cnt = Student.objects.filter(Q(batch = batch), Q(rollno__startswith = q) | Q(name__icontains = q) | Q(user__email__icontains = q) ).distinct().order_by('id').count()
                  else:
                        students_list = Student.objects.filter(batch = batch).order_by('id').all()
                        total_cnt = Student.objects.filter(batch = batch).order_by('id').count()
            
            p = Paginator(students_list, students_per_page)
            
            page_number = request.GET.get('page')
            if page_number is None:
                  page_number = 1
            page_number = int(page_number)
            total_pages = p.num_pages
            if (page_number>total_pages or page_number<1):
                  return Response({'error':'Page does not exist'}, status=status.HTTP_400_BAD_REQUEST)
            
            students = p.page(page_number)
            serializer = StudentProfileSerializer(students, many=True, context={'is_admin': True})
            
            return Response({'status':'success', 'data':serializer.data, 'total_pages':total_pages, 'total_entries': total_cnt}, status=status.HTTP_200_OK)


class getGroups(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]
      
      def get(self, request):
            q = request.GET.get('q')
            groups_per_page = request.GET.get('groups_per_page')
            
            if q is not None and q.strip()!='':
                  groups_list = Group.objects.filter(Q(leader__rollno__startswith = q) | Q(leader__name__icontains = q) | Q(leader__user__email__icontains = q) | Q(members__rollno__startswith = q) | Q(members__name__icontains = q) | Q(members__user__email__icontains = q) ).distinct().select_related('leader').prefetch_related('members').order_by('id').all()
                  total_cnt = Group.objects.filter(Q(leader__rollno__startswith = q) | Q(leader__name__icontains = q) | Q(leader__user__email__icontains = q) | Q(members__rollno__startswith = q) | Q(members__name__icontains = q) | Q(members__user__email__icontains = q) ).distinct().count()
            else:
                  groups_list = Group.objects.select_related('leader').prefetch_related('members').order_by('id').all()
                  total_cnt = Group.objects.count()
            p = Paginator(groups_list, groups_per_page)
            
            page_number = request.GET.get('page')
            if page_number is None:
                  page_number = 1
            page_number = int(page_number)
            total_pages = p.num_pages
            if (page_number>total_pages or page_number<1):
                  return Response({'error':'Page does not exist'}, status=status.HTTP_400_BAD_REQUEST)
            
            groups = p.page(page_number)
            serializer = GroupSerializer(groups, many=True)
            
            return Response({'status':'success', 'data':serializer.data, 'total_pages':total_pages, 'total_entries': total_cnt}, status=status.HTTP_200_OK)


class getGroup(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]
      
      def get(self, request):
            id = request.data.get('id')
            group = Group.objects.filter(id=id).first()
            if group is None:
                  return Response({'error':'Group does not exist'}, status=status.HTTP_404_NOT_FOUND)
            serializer = GroupDetailSerializer(group)
            return Response({'status':'success', 'data':serializer.data}, status=status.HTTP_200_OK)


class getDefaulters(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def get(self, request):
            q = request.GET.get('q')
            defaulters_per_page = request.GET.get('defaulters_per_page')
            
            if q is not None:
                  defaulters_list = Defaulter.objects.filter(Q(student__rollno__startswith = q) | Q(student__name__contains = q) | Q(student__user__email__contains = q) )
                  total_cnt = Defaulter.objects.filter(Q(student__rollno__startswith = q) | Q(student__name__contains = q) | Q(student__user__email__contains = q) ).count()
            else:
                  defaulters_list = Defaulter.objects.all()
                  total_cnt = Defaulter.objects.count()
            p = Paginator(defaulters_list, defaulters_per_page)
            
            page_number = request.GET.get('page')
            if page_number is None:
                  page_number = 1
            page_number = int(page_number)
            total_pages = p.num_pages
            if (page_number>total_pages or page_number<1):
                  return Response({'error':'Page does not exist'}, status=status.HTTP_400_BAD_REQUEST)
            
            defaulters = p.page(page_number)
            serializer = DefaulterSerializer(defaulters, many=True)
            
            return Response({'status':'success', 'data':serializer.data, 'total_pages':total_pages, 'total_entries': total_cnt}, status=status.HTTP_200_OK)


class ProfileView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def get(self, request):
            user = request.user
            serializer = ProfileSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)


class createFAQ(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]
      
      def post(self, request):
            question = request.data.get('question')
            answer = request.data.get('answer')
            faq = Faq(question=question, answer=answer)
            faq.save()
            return Response(status=status.HTTP_200_OK)
     

class getFAQ(APIView):
      permission_classes = [IsAuthenticated]
      
      def get(self, request):
            faqs = Faq.objects.all()
            serializer = FAQSerializer(faqs, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

  
class deleteFAQ(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]
      
      def post(self, request):
            id = request.data.get('id')
            faq = Faq.objects.filter(id=id).first()
            if faq is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)
            faq.delete()
            return Response(status=status.HTTP_200_OK)
      
class updateFAQ(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]
      
      def post(self, request):
            id = request.data.get('id')
            faq = Faq.objects.filter(id=id).first()
            if faq is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)
            question = request.data.get('question')
            answer = request.data.get('answer')
            
            if question is not None:
                  faq.question = question
            if answer is not None:
                  faq.answer = answer
            faq.save()
            return Response(status=status.HTTP_200_OK)
            

class UpdateSectionsAllotmentStatusView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            data = request.data
            for item in data:
                  if item.get('id') is None or item.get('is_allotment_enabled') is None:
                        return Response(status=status.HTTP_400_BAD_REQUEST)
                  instance = Section.objects.filter(id=item.get('id')).first()
                  if instance is None:
                        return Response(status=status.HTTP_400_BAD_REQUEST)
                  instance.is_allotment_enabled = item.get('is_allotment_enabled', instance.is_allotment_enabled)
                  instance.save()
            return Response(status=status.HTTP_200_OK)


class sendReminderMail(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]
      
      def post(self, request):
            section_ids = request.data.get('sections')
            if section_ids is None:
                  return Response({'detail': 'Invalid data!'}, status=status.HTTP_400_BAD_REQUEST)
            
            for section_id in section_ids:
                  section = Section.objects.filter(id=section_id).select_related('batch').first()
                  if section is None or not section.is_allotment_enabled:
                        continue
                  groups = Group.objects.filter(
                        Q(leader__batch = section.batch) 
                        & Q(leader__gender = section.gender)
                        & Q(is_retained = False)
                        & Q(is_preferences_filled = False)
                  ).select_related('leader').prefetch_related('members').all()
                  for group in groups:
                        for student in group.members.all():
                              send_reminder_mail.delay(student.name, student.user.email)
                        send_reminder_mail.delay(group.leader.name, group.leader.user.email)
            
            return Response(status=status.HTTP_200_OK)


class CreateGroupView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            leader_rollno = request.data.get('leader')
            leader = Student.objects.filter(rollno=leader_rollno).select_related('leader_of_group', 'defaulter', 'group').first()
            if leader is None:
                  return Response({'detail': f'No student found with roll number {leader_rollno}!'}, status=status.HTTP_403_FORBIDDEN)
            
            group = None
            left_group_mails = []
            members_list = []
            member_exist = {}

            try:
                  _ = leader.defaulter
                  return Response({'detail': f'{leader_rollno} is currently suspended from Hostel Allocation Process!'}, status=status.HTTP_403_FORBIDDEN)
            except ObjectDoesNotExist:
                  pass

            try:
                  gp = leader.leader_of_group
                  if gp.members.count()>0:
                        return Response({'detail': f'{leader_rollno} is a leader of another group! Please first transfer ownership of that group'}, status=status.HTTP_403_FORBIDDEN)
                  group = gp
            except ObjectDoesNotExist:
                  pass
            
            member_exist[leader_rollno] = True
            members_data = request.data.get('members')

            for member_rollno in members_data:
                  if member_exist.get(member_rollno, False):
                        return Response({'detail': f'{member_rollno} is present multiple times in request data!'}, status=status.HTTP_403_FORBIDDEN)

                  member = Student.objects.filter(rollno=member_rollno).select_related('leader_of_group', 'defaulter', 'group').first()
                  if member is None:
                        return Response({'detail': f'No student found with roll number {member_rollno}!'}, status=status.HTTP_403_FORBIDDEN)
                  
                  try:
                        _ = member.defaulter
                        return Response({'detail': f'{member_rollno} is currently suspended from Hostel Allocation Process!'}, status=status.HTTP_403_FORBIDDEN)
                  except ObjectDoesNotExist:
                        pass
                  
                  try:
                        gp = member.leader_of_group
                        if gp.members.count()>0:
                              return Response({'detail': f'{member_rollno} is a leader of another group! Please first transfer the ownership of thet group'}, status=status.HTTP_403_FORBIDDEN)
                  except ObjectDoesNotExist:
                        pass

                  members_list.append(member)
                  member_exist[member_rollno] = True
            
            with transaction.atomic():
                  if leader.group is not None:
                        left_group_mails.append((leader, leader.group))
                        leader.group = None
                        group = Group.objects.create(leader=leader, cg=leader.cg)
                  
                  cg = leader.cg
                  if group is None:
                        group = Group.objects.create(leader=leader, cg=cg)

                  for member in members_list:            
                        try:
                              gp = member.leader_of_group
                              if gp.members.count()==0:
                                    gp.delete()
                        except ObjectDoesNotExist:
                              if member.group is not None:
                                    left_group_mails.append((member, member.group))

                        member.group = group
                        cg += member.cg
                        member.save()

                  cg = round(cg / (len(members_data) + 1), 2)
                  group.cg = cg
                  group.save()

            for target in left_group_mails:
                  left_group_mail.delay(target[1].leader.name, target[0].name, target[0].rollno, target[1].leader.user.email)
                  target_ms = target[1].members.all()
                  for m in target_ms:
                        left_group_mail.delay(m.name, target[0].name, target[0].rollno, m.user.email)
            
            for m in members_list:
                  joined_group_mail.delay(leader.name, leader.user.email, leader.rollno, m.name, m.user.email)
                  joined_group_to_members.delay(leader.name, m.name, m.rollno, group.leader.user.email)
                  for m2 in members_list:
                        if m2 != m:
                              joined_group_to_members.delay(m2.name, m.name, m.rollno, m2.user.email)

            serializer = GroupDetailSerializer(group)
            return Response(serializer.data, status=status.HTTP_200_OK)


# Edit Existing group routes

class AddToGroupView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(sef, request):
            student = Student.objects.filter(rollno=request.data.get('rollno')).select_related('leader_of_group', 'defaulter', 'group').prefetch_related('group__members').first()
            if student is None:
                  return Response({'detail': 'No Student Found!'}, status=status.HTTP_404_NOT_FOUND)
            
            group = Group.objects.filter(id=request.data.get('group')).prefetch_related('members').first()
            if group is None:
                  return Response({'detail': 'No Group Found!'}, status=status.HTTP_404_NOT_FOUND)
            
            try:
                  _ = student.defaulter
                  return Response({"detail": "This student is currently suspended from Hostel Allocation Process"}, status=status.HTTP_403_FORBIDDEN)
            except ObjectDoesNotExist:
                  pass

            with transaction.atomic():
                  try:
                        gp = student.leader_of_group
                        if gp==group:
                              return Response({'detail': 'This student is already leader of this group.'}, status=status.HTTP_403_FORBIDDEN)
                        if gp.members.count()>0:
                              return Response({'detail': 'This student is leader of some another group! First transfer ownership of that group.'}, status=status.HTTP_403_FORBIDDEN)
                        gp.delete()
                  except ObjectDoesNotExist:
                        if student.group is not None:
                              if student.group==group:
                                    return Response({'detail': 'This student is already a member of this group.'}, status=status.HTTP_403_FORBIDDEN)
                              
                              gp = student.group
                              left_group_mail.delay(gp.leader.name, student.name, student.rollno, gp.leader.user.email)
                              ms = gp.members.all()
                              for m in ms:
                                    if m != student:
                                          left_group_mail.delay(m.name, student.name, student.rollno, m.user.email)
                              
                              # due to prefetch, members were fetched before execution of above statement, so adjust cg of removed student
                              updatedcg = gp.leader.cg  - student.cg
                              cnt = 0
                              for member in ms:
                                    updatedcg += member.cg
                                    cnt += 1
                              updatedcg /= cnt
                              gp.cg = round(updatedcg, 2)
                              gp.save()

                  student.group = group
                  student.save()

                  # due to prefetch, members were fetched before execution of above statement, so adjust cg of added student
                  updatedcg = group.leader.cg + student.cg
                  cnt = 2
                  for member in group.members.all():
                        updatedcg += member.cg
                        cnt += 1
                  updatedcg /= cnt
                  group.cg = round(updatedcg, 2)
                  group.save()

            lead = group.leader
            joined_group_mail.delay(lead.name, lead.user.email, lead.rollno, student.name, student.user.email)
            joined_group_to_members.delay(lead.name, student.name, student.rollno, group.leader.user.email)
            members = group.members.all()
            for member in members:
                  joined_group_to_members.delay(member.name, student.name, student.rollno, member.user.email)

            serializer = GroupDetailSerializer(group)
            return Response(serializer.data, status=status.HTTP_200_OK)


class ChangeGroupLeaderView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            newleader = Student.objects.filter(rollno=request.data.get('rollno')).select_related('leader_of_group', 'defaulter').first()
            if newleader is None:
                  return Response({'detail': 'No Student Found!'}, status=status.HTTP_404_NOT_FOUND)
            
            group = Group.objects.filter(id=request.data.get('group')).select_related('leader').prefetch_related('members').first()
            if group is None:
                  return Response({'detail': 'No Group Found!'}, status=status.HTTP_404_NOT_FOUND)
            
            if group.leader==newleader:
                  return Response({'detail': 'This student already the leader of this group.'}, status=status.HTTP_403_FORBIDDEN)
            
            if newleader.group is None or newleader.group!=group:
                  return Response({'detail': 'You can only tranfer ownership to one of the group members!'}, status=status.HTTP_403_FORBIDDEN)

            with transaction.atomic():
                  newleader.group = None
                  newleader.save()

                  old_leader = group.leader
                  group.leader = newleader
                  group.save()
                  
                  old_leader.group = group
                  old_leader.save()
            
            members = group.members.all()
            for member in members:
                  send_teamleader_change_mail.delay(group.leader.name, group.leader.rollno, member.user.email, member.name)
            send_teamleader_change_mail.delay(group.leader.name, group.leader.rollno, group.leader.user.email, group.leader.name)
            
            serializer = GroupDetailSerializer(group)
            return Response(serializer.data, status=status.HTTP_200_OK)


class RemoveFromGroupView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            student = Student.objects.filter(rollno=request.data.get('rollno')).select_related('leader_of_group', 'defaulter', 'group').prefetch_related('group__members').first()
            if student is None:
                  return Response({'detail': 'No Student Found!'}, status=status.HTTP_404_NOT_FOUND)
            
            try:
                  _ = student.leader_of_group
                  return Response({'detail': 'You must tranfer the group ownership to one of the group members to leave this group!'}, status=status.HTTP_403_FORBIDDEN)
            except:
                  group = student.group
                  if group is None:
                        return Response({'detail': 'You can only remove a group member from the group!'}, status=status.HTTP_403_FORBIDDEN)
            
            with transaction.atomic():
                  student.group = None
                  student.save()

                  # due to prefetch, members were fetched before execution of above statement, so adjust cg of removed student
                  updatedcg = group.leader.cg - student.cg
                  cnt = 0
                  for member in group.members.all():
                        updatedcg += member.cg
                        cnt += 1
                  updatedcg /= cnt
                  group.cg = round(updatedcg, 2)
                  group.save()

                  indvgroup = Group(leader = student, cg = student.cg)
                  indvgroup.save()
            
            left_group_mail.delay(group.leader.name, student.name, student.rollno, group.leader.user.email)
            
            members = group.members.all()
            for member in members:
                  left_group_mail.delay(member.name, student.name, student.rollno, member.user.email)
            
            serializer = GroupDetailSerializer(group)
            return Response(serializer.data, status=status.HTTP_200_OK)


class DeleteGroupsView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def delete(self, request):
            ids = request.data.get('ids')
            if not isinstance(ids, list):
                  return Response({'detail': 'Invalid data!'}, status=status.HTTP_400_BAD_REQUEST)
            with transaction.atomic():
                  for id in ids:
                        group = Group.objects.filter(id=id).prefetch_related('members', 'preferences').first()
                        if group is None:
                              continue
                        for member in group.members.all():
                              member.group = None
                              newgp = Group.objects.create(leader=member, cg=member.cg)
                              member.save()
                        for preference in group.preferences.all():
                              preference.delete()
            
                        # left_group_mail.delay(group.leader.name, student.name, student.rollno, group.leader.user.email)
                        
                        # members = group.members.all()
                        # for member in members:
                        #       left_group_mail.delay(member.name, student.name, student.rollno, member.user.email)

            return Response(status=status.HTTP_200_OK)


class ImportStudentsView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def post(self, request):
            file = request.data.get('file')

            if file is None:
                  return Response({'detail': 'File Not Found!'}, status=status.HTTP_400_BAD_REQUEST)
            
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M')}_{file.name}"

            if filename.split('.')[-1]!='xlsx':
                  return Response({'detail': 'Only xlsx file format supported!'}, status=status.HTTP_400_BAD_REQUEST)
            
            storage = default_storage
            storage.location = os.path.join(settings.BASE_DIR, 'imported-data', 'student')
            filename = storage.save(filename, file)

            path = os.path.join(settings.BASE_DIR, 'imported-data', 'student', filename)
            wb = load_workbook(path)
            ws = wb.active

            # Check if all fields are present in given data
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=11):
                  for i in range(len(row)):
                        if row[i].value!=studentfileformat[i]:
                              return Response({
                                    'detail': f"Invalid File Format. Found {row[i].value}, but required was {studentfileformat[i]}, Correct File Format is {str(studentfileformat)}"
                              }, status=status.HTTP_400_BAD_REQUEST)

            import_students.delay(filename)

            return Response(status=status.HTTP_202_ACCEPTED)


class ImportDefaultersView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            file = request.data.get('file')
            
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M')}_{file.name}"

            if filename.split('.')[-1]!='xlsx':
                  return Response({'error': 'Only xlsx file format supported!'}, status=status.HTTP_400_BAD_REQUEST)
            
            storage = default_storage
            storage.location = os.path.join(settings.BASE_DIR, 'imported-data', 'defaulter')
            filename = storage.save(filename, file)
            
            path = os.path.join(settings.BASE_DIR, 'imported-data', 'defaulter', filename)
            wb = load_workbook(path)
            ws = wb.active

            # check if required fields are present in data
            if ws['A1'].value!='student':
                  return Response({
                        'detail': f"'student' field missing! Required 'student' but found '{ws['A1'].value}' at cell A1."
                  }, status=status.HTTP_400_BAD_REQUEST)
            
            import_defaulters.delay(filename)
            
            return Response(status=status.HTTP_202_ACCEPTED)


class ExportGroupsView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            if request.data.get('section')=='all':
                  sections = Section.objects.select_related('batch').all()
                  filename = f"export/group/all_{''.join(choice(string.ascii_letters + string.digits) for _ in range(40))}.xlsx"
            else:
                  section = Section.objects.filter(id=request.data.get('section')).select_related('batch').first()
                  if section is None:
                        return Response(status=status.HTTP_400_BAD_REQUEST)
                  filename = f"export/group/{section.batch.name}_{section.gender}_{''.join(choice(string.ascii_letters + string.digits) for _ in range(40))}.xlsx"
                  sections = [section]
            
            include = OrderedDict([
                  ('Roll Number', request.data.get('include_rollno', True)),
                  ('Email Address', request.data.get('include_email', True)),
                  ('Name', request.data.get('include_name', True)),
                  ('Phone Number', request.data.get('include_phoneno', True)),
                  ('CG', request.data.get('include_cg', True)),
                  ('Gender', request.data.get('include_gender', True)),
                  ('Batch', request.data.get('include_batch', True)),
                  ('Current Hostel', request.data.get('include_current_hostel', True)),
                  ('Current Room Type', request.data.get('include_current_hostel', True)),
                  ('Preview Hostel', request.data.get('include_preview_hostel', True)),
                  ('Preview Room Type', request.data.get('include_preview_hostel', True)),
                  ('Alloted Hostel', request.data.get('include_alloted_hostel', True)),
                  ('Alloted Room Type', request.data.get('include_alloted_hostel', True)),
            ])
            
            path = os.path.join(settings.MEDIA_ROOT, filename)

            wb = Workbook()
            ws = wb.active
            
            header = ['Group ID', 'Group Size', 'Group CG']
            student_fields = []
            for key in include.keys():
                  if include[key]:
                        student_fields.append(key)
            
            preferences_cnt = 0
            group_size_limit = 1
            for section in sections:
                  preferences_cnt = max(preferences_cnt, section.choices.count())
                  group_size_limit = max(group_size_limit, section.group_size_limit)
            
            if group_size_limit==1:
                  header_student_stakeholders = ['Student']
            else:
                  header_student_stakeholders = ['Leader']
                  for i in range(1, group_size_limit):
                        header_student_stakeholders.append(f'Member {i}')
            for member_name in header_student_stakeholders:
                  for field in student_fields:
                        header.append(f"{member_name} {field}")

            header.append('Is Retaining')

            for i in range(1, preferences_cnt+1):
                  header.append(f'Preference {i} Hostel')            
                  header.append(f'Preference {i} Room Type')
                  
            ws.append(header)
            for col in ws.iter_cols(min_col=1, max_col=len(header), min_row=1, max_row=1):
                  col[0].font = Font(bold=True)
            ws.freeze_panes = 'A2'
            
            for section in sections:
                  queryset = Group.objects.filter(leader__batch=section.batch, leader__gender=section.gender).select_related('leader__user', 'leader__batch', 'leader__current_room__hostel', 'leader__alloted_room__hostel').prefetch_related('members__user', 'members__batch', 'members__current_room__hostel', 'members__alloted_room__hostel', 'preferences__room_type_choice__room_type__hostel').all()

                  for group in queryset:
                        cnt = 0
                        row = [group.id, group.members.count() + 1, group.cg]
                        members_list = [group.leader] + list(group.members.all())
                        for member in members_list:
                              if include.get('Roll Number', False):
                                    row.append(member.rollno)
                              if include.get('Email Address', False):
                                    row.append(member.user.email)
                              if include.get('Name', False):
                                    row.append(member.name)
                              if include.get('Phone Number', False):
                                    row.append(member.phoneno)
                              if include.get('CG', False):
                                    row.append(member.cg)
                              if include.get('Gender', False):
                                    row.append(member.gender)
                              if include.get('Batch', False):
                                    row.append(member.batch.name)
                              if include.get('Current Hostel', False):
                                    if member.current_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.current_room.hostel.name)
                              if include.get('Current Room Type', False):
                                    if member.current_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.current_room.name)
                              if include.get('Preview Hostel', False):
                                    if member.preview_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.preview_room.hostel.name)
                              if include.get('Preview Room Type', False):
                                    if member.preview_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.preview_room.name)
                              if include.get('Alloted Hostel', False):
                                    if member.alloted_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.alloted_room.hostel.name)
                              if include.get('Alloted Room Type', False):
                                    if member.alloted_room is None:
                                          row.append('')
                                    else:
                                          row.append(member.alloted_room.name)
                              cnt += 1

                        while cnt < group_size_limit:
                              if include.get('Roll Number', False):
                                    row.append('')
                              if include.get('Email Address', False):
                                    row.append('')
                              if include.get('Name', False):
                                    row.append('')
                              if include.get('Phone Number', False):
                                    row.append('')
                              if include.get('CG', False):
                                    row.append('')
                              if include.get('Gender', False):
                                    row.append('')
                              if include.get('Batch', False):
                                    row.append('')
                              if include.get('Current Hostel', False):
                                    row.append('')
                              if include.get('Current Room Type', False):
                                    row.append('')
                              if include.get('Preview Hostel', False):
                                    row.append('')
                              if include.get('Preview Room Type', False):
                                    row.append('')
                              if include.get('Alloted Hostel', False):
                                    row.append('')
                              if include.get('Alloted Room Type', False):
                                    row.append('')
                              cnt += 1

                        if group.is_retained:
                              row.append(True)
                        else:
                              row.append(False)

                        for preference in group.preferences.order_by('priority').all():
                              row.append(preference.room_type_choice.room_type.hostel.name)
                              row.append(preference.room_type_choice.room_type.name)
                        ws.append(row)
                  
            wb.save(path)
            return Response({
                  'link': f"{settings.MEDIA_URL}{filename}"
            }, status=status.HTTP_200_OK)


class ExportStudentsView(APIView):
      permission_classes = [IsAuthenticated & IsAdmin]

      def post(self, request):
            if request.data.get('batch')=='all':
                  batches = Batch.objects.prefetch_related('students__user', 'students__batch', 'students__current_room__hostel', 'students__alloted_room__hostel').all()
                  filename = f"export/student/batch_all_{''.join(choice(string.ascii_letters + string.digits) for _ in range(40))}.xlsx"
            else:
                  batch = Batch.objects.filter(id=request.data.get('batch')).prefetch_related('students__user', 'students__batch', 'students__current_room__hostel', 'students__alloted_room__hostel').first()
                  if batch is None:
                        return Response(status=status.HTTP_400_BAD_REQUEST)
                  filename = f"export/student/batch_{batch.id}_{''.join(choice(string.ascii_letters + string.digits) for _ in range(40))}.xlsx"
                  batches = [batch]
            
            path = os.path.join(settings.MEDIA_ROOT, filename)
            wb = Workbook()
            ws = wb.active

            include = OrderedDict([
                  ('Roll Number', request.data.get('include_rollno', True)),
                  ('Email Address', request.data.get('include_email', True)),
                  ('Name', request.data.get('include_name', True)),
                  ('Phone Number', request.data.get('include_phoneno', True)),
                  ('CG', request.data.get('include_cg', True)),
                  ('Gender', request.data.get('include_gender', True)),
                  ('Batch', request.data.get('include_batch', True)),
                  ('Current Hostel', request.data.get('include_current_hostel', True)),
                  ('Current Room Type', request.data.get('include_current_hostel', True)),
                  ('Preview Hostel', request.data.get('include_preview_hostel', True)),
                  ('Preview Room Type', request.data.get('include_preview_hostel', True)),
                  ('Alloted Hostel', request.data.get('include_alloted_hostel', True)),
                  ('Alloted Room Type', request.data.get('include_alloted_hostel', True)),
            ])

            header = []
            for key in include.keys():
                  if include[key]:
                        header.append(key)
            
            ws.append(header)
            for col in ws.iter_cols(min_col=1, max_col=len(header), min_row=1, max_row=1):
                  col[0].font = Font(bold=True)
            ws.freeze_panes = 'A2'

            for batch in batches:
                  queryset = batch.students.all()

                  for student in queryset:
                        row = []
                        if include.get('Roll Number', False):
                              row.append(student.rollno)
                        if include.get('Email Address', False):
                              row.append(student.user.email)
                        if include.get('Name', False):
                              row.append(student.name)
                        if include.get('Phone Number', False):
                              row.append(student.phoneno)
                        if include.get('CG', False):
                              row.append(student.cg)
                        if include.get('Gender', False):
                              row.append(student.gender)
                        if include.get('Batch', False):
                              row.append(student.batch.name)
                        if include.get('Current Hostel', False):
                              if student.current_room is None:
                                    row.append('')
                              else:
                                    row.append(student.current_room.hostel.name)
                        if include.get('Current Room Type', False):
                              if student.current_room is None:
                                    row.append('')
                              else:
                                    row.append(student.current_room.name)
                        if include.get('Preview Hostel', False):
                              if student.preview_room is None:
                                    row.append('')
                              else:
                                    row.append(student.preview_room.hostel.name)
                        if include.get('Preview Room Type', False):
                              if student.preview_room is None:
                                    row.append('')
                              else:
                                    row.append(student.preview_room.name)
                        if include.get('Alloted Hostel', False):
                              if student.alloted_room is None:
                                    row.append('')
                              else:
                                    row.append(student.alloted_room.hostel.name)
                        if include.get('Alloted Room Type', False):
                              if student.alloted_room is None:
                                    row.append('')
                              else:
                                    row.append(student.alloted_room.name)
                        ws.append(row)

            wb.save(path)

            return Response({
                  'link': f"{settings.MEDIA_URL}{filename}"
            }, status=status.HTTP_200_OK)


class ExportDefaultersView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def get(self, request):
            filename = f"export/defaulter/{''.join(choice(string.ascii_letters + string.digits) for _ in range(40))}.xlsx"
            path = os.path.join(settings.MEDIA_ROOT, filename)
            queryset = Defaulter.objects.select_related('student__user', 'student__batch', 'student__current_room__hostel', 'student__alloted_room__hostel').all()
            
            wb = Workbook()
            ws = wb.active

            ws.append(['Roll Number', 'Email Address', 'Name', 'Phone Number', 'CG', 'Batch', 'Gender', 'Current Hostel', 'Current Room Type', 'Alloted Hostel', 'Alloted Room Type'])
            for col in ws.iter_cols(min_col=1, max_col=11, min_row=1, max_row=1):
                  col[0].font = Font(bold=True)
            ws.freeze_panes = 'A2'

            for defaulter in queryset:
                  student = defaulter.student
                  row = [student.rollno, student.user.email, student.name, student.phoneno, student.cg, student.batch.name, 'Female' if student.gender=='F' else 'Male']
                  if student.current_room is not None:
                        row.append(student.current_room.hostel.name)
                        row.append(student.current_room.name)
                  else:
                        row.append('')
                        row.append('')
                  if student.alloted_room is not None:
                        row.append(student.alloted_room.hostel.name)
                        row.append(student.alloted_room.name)
                  else:
                        row.append('')
                        row.append('')
                  ws.append(row)

            wb.save(path)
            return Response({
                  'link': f"{settings.MEDIA_URL}{filename}"
            }, status=status.HTTP_200_OK)


class AllotmentView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            section_ids = request.data.get('sections')
            allot.delay(section_ids)
            return Response(status=status.HTTP_202_ACCEPTED)


class BatchAnalyticsView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def get(self, request):
            batch_id = request.GET.get('batch')
            if batch_id is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)
            
            batch_id = batch_id.strip()
            if not batch_id.isnumeric():
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            
            batch = Batch.objects.filter(id=batch_id).first()
            if batch is None:
                  return Response(status=status.HTTP_404_NOT_FOUND)

            total_students_cnt = batch.students.count()
            male_students_cnt = batch.students.filter(gender='M').count()
            female_students_cnt = batch.students.filter(gender='F').count()

            total_preferences_filled = Group.objects.filter(leader__batch=batch).annotate(
                  preferences_count = Count('preferences'),
                  members_count = Count('members') + 1
            ).filter(preferences_count__gt=0).aggregate(Sum('members_count', default=0))
            male_preferences_filled = Group.objects.filter(leader__batch=batch, leader__gender='M').annotate(
                  preferences_count = Count('preferences'),
                  members_count = Count('members') + 1
            ).filter(preferences_count__gt=0).aggregate(Sum('members_count', default=0))
            female_preferences_filled = Group.objects.filter(leader__batch=batch, leader__gender='F').annotate(
                  preferences_count = Count('preferences'),
                  members_count = Count('members') + 1
            ).filter(preferences_count__gt=0).aggregate(Sum('members_count', default=0))

            choice_analytics = {}

            for gender in ['M', 'F']:
                  section = Section.objects.filter(batch=batch, gender=gender).first()
                  if section is None:
                        choice_analytics[gender] = None
                  else:
                        choice_analytics[gender] = {}
                        choices = section.choices.all()
                        for choice in choices:
                              choice_analytics[gender][choice.room_type.name] = {}
                              for p in range(1, section.choices.count()+1):
                                    choice_priority_students = Group.objects.filter(leader__batch=batch, leader__gender=gender, preferences__priority=p, preferences__room_type_choice=choice).annotate(
                                          members_count = Count('members') + 1
                                    ).aggregate(Sum('members_count', default=0))
                                    choice_analytics[gender][choice.room_type.name][p] = choice_priority_students.get('members_count__sum', 0)

            res = {
                  'total_students': total_students_cnt,
                  'male_students': male_students_cnt,
                  'female_students': female_students_cnt,
                  'total_preferences_filled': total_preferences_filled.get('members_count__sum', 0),
                  'male_preferences_filled': male_preferences_filled.get('members_count__sum', 0),
                  'female_preferences_filled': female_preferences_filled.get('members_count__sum', 0),
                  'choice_analytics': choice_analytics
            }

            return Response(res, status=status.HTTP_200_OK)
      
      

class CannotBeAllocatedRoomView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            hostel = request.data.get('hostel')
             
            start = int(request.data.get('start'))
            end = int(request.data.get('end'))
            if start is None or end is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            rooms = Room.objects.filter(room_no__gte=start, room_no__lte=end, room_type__hostel = hostel).all()
            for room in rooms:
                  if room.is_allotted:
                        return Response({'Error':f'Students are currently residing in room number {room.room_no}'},status=status.HTTP_400_BAD_REQUEST)
                  room.is_registered = False
                  room.save()
            return Response(status=status.HTTP_200_OK)
      
class CanBeAllocatedRoomView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            hostel = request.data.get('hostel')
             
            start = int(request.data.get('start'))
            end = int(request.data.get('end'))
            if start is None or end is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            rooms = Room.objects.filter(room_no__gte=start, room_no__lte=end, room_type__hostel = hostel).all()
            for room in rooms:
                  room.is_registered = True
                  room.save()
            return Response(status=status.HTTP_200_OK)
      
class AddRoomRangeView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            room_type = request.data.get('room_type')
            level = request.data.get('level')
            if room_type is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            
            start = int(request.data.get('start'))
            end = int(request.data.get('end'))
            if start is None or end is None or level is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            room_type = RoomType.objects.filter(room_type = room_type).first()
            if Room.objects.filter(room_no__gte=start, room_no__lte=end, room_type = room_type).exists():
                  return Response({'Error':'Room already exists'},status=status.HTTP_400_BAD_REQUEST)
            for i in range(start,end+1):
                  room = Room(room_no = i, room_type = room_type, level = level)
                  room.save()
            return Response(status=status.HTTP_200_OK)
      

class DeleteRoomRangeView(APIView):
      permission_classes = [IsAuthenticated, IsAdmin]

      def post(self, request):
            room_type = request.data.get('room_type')
            if room_type is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            
            start = int(request.data.get('start'))
            end = int(request.data.get('end'))
            if start is None or end is None:
                  return Response(status=status.HTTP_400_BAD_REQUEST)
            room_type = RoomType.objects.filter(room_type = room_type).first()
            rooms = Room.objects.filter(room_no__gte=start, room_no__lte=end, room_type = room_type).all()
            for room in rooms:
                  room.delete()
            return Response(status=status.HTTP_200_OK)