from __future__ import absolute_import, unicode_literals

from django.db.models import Q, ObjectDoesNotExist, Count
from django.core.mail import send_mail, get_connection, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.core.cache import cache
from django.db import transaction

from config.celery import app

from student.models import Student, Group, Defaulter, Batch, Section
from user.models import User
from preference.models import Hostel, RoomType, Preference
from dashboard.models import AllotmentLogsStudent, AllotmentLogsGroup, AllotmentStatus, AcademicSession
from student.tasks import left_group_mail, send_teamleader_change_mail

from datetime import datetime
import os, string, json
from openpyxl import load_workbook
from random import choice


@app.task(name = "send_reminder_mail")
def send_reminder_mail(name, email):
      subject = "Hi " + name +", Reminder for filling Hostel Preferences"
      
      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()
      context = {
            "name" : name
      }
      html_message = render_to_string('dashboard/remindermail.html', context)
      msg = strip_tags(html_message)
      email = EmailMultiAlternatives(subject, msg, settings.EMAIL_HOST_USERS[idx], [email], reply_to = ['queries_studentaffairs@thapar.edu'], connection=connection)
      email.attach_alternative(html_message, "text/html")
      email.send()
      
      # send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (email, ), html_message=html_message, connection=connection, fail_silently=False)
      connection.close()
      
      return f"\nReminder mail sent to {email}\n"


@app.task(name = "send_start_allocation_mail")
def start_preference_filling(section_id):
      section = Section.objects.filter(id=section_id).first()

      groups = Group.objects.filter(leader__batch=section.batch, gender=section.gender).all()

      for group in groups:
            students_list = [group.leader]
            for member in group.members.all():
                  students_list.append(member)
            
            for student in students_list:
                  user = student.user
                  slug_instance = ResetSlug.objects.filter(user=user).first()
                  if slug_instance is None:
                        while True:
                              slug = ''.join(choice(string.ascii_letters + string.digits + '-') for _ in range(135))
                              if not ResetSlug.objects.filter(slug=slug).exists():
                                    break
                        slug_instance = ResetSlug(user=user, slug=slug)
                        slug_instance.save()
                  user_slug = slug_instance.slug

                  subject = f"Hi {student.name}, Hostel Preference Filling Process has Started"
                  
                  idx = cache.get('emailIdIndex', 0)
                  cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
                  
                  context = {
                        "name" : student.name,
                        "url": "https://allotment.onlinehostel.in/auth/forgot-password/" + user_slug,
                  }
                  html_message = render_to_string('dashboard/startallocationmail.html', context)
                  msg = strip_tags(html_message)
                  
                  connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
                  connection.open()
                  
                  email = EmailMultiAlternatives(subject, msg, settings.EMAIL_HOST_USERS[idx], [user.email], reply_to = ['queries_studentaffairs@thapar.edu'], connection=connection)
                  email.attach_alternative(html_message, "text/html")
                  email.send()
                  
                  connection.close()
      
      subject = f'Execution results ready for the task Start Preference Filling'
      msg = f"Successfully started Preferences Filling for section {section.batch.name} {'Girls' if section.gender=='F' else 'Boys'} and sent mails regarding the same."

      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()

      send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL), connection=connection, fail_silently=False)

      connection.close()

      return 1


studentfileformat = ('rollno', 'email', 'name', 'phoneno', 'cg', 'batch', 'gender', 'current_hostel', 'current_room_type', 'alloted_hostel', 'alloted_room_type')

@app.task(name='import_students', queue='hms-admin-task')
def import_students(filename):
      start_time = datetime.now()

      path = os.path.join(settings.BASE_DIR, 'imported-data', 'student', filename)
      wb = load_workbook(path, data_only=True)
      ws = wb.active

      log_file_path = os.path.join(settings.LOGS_ROOT, 'import_students.log')
      with open(log_file_path, 'a') as f:
            f.write(f"\nProcessing file {filename} at {start_time}...\n")

      # Check if all fields are present in given data
      for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=11):
            for i in range(len(row)):
                  if row[i].value!=studentfileformat[i]:
                        with open(log_file_path, 'a') as f:
                              f.write(f"Invalid File Format. Found {row[i].value}, but required was {studentfileformat[i]}\n")
                              f.write(f"Correct File Format is {str(studentfileformat)}\n")
                              f.write("Aborting...\n\n")
                              f.write("\n")
                        raise Exception(f"Found {row[i]}, but required was {studentfileformat[i]}\n")

      successCnt = 0
      failureCnt = 0
      cnt = 1
      errors = []

      for row in ws.iter_rows(min_row=2, min_col=1, max_col=11):
            try:
                  with transaction.atomic():
                        student = None
                        
                        # Extract data into variables and data preprocessing
                        phoneno = row[3].value

                        cg = row[4].value
                        if cg is not None:
                              cg = round(float(cg), 2)

                        batch = row[5].value
                        if batch is not None:
                              batch = batch.strip()
                              batch = Batch.objects.filter(name__iexact = batch).first()
                              if batch is None:
                                    batch = Batch(name = row[5].value.strip())
                                    batch.save()
                        
                        rollno = row[0].value
                        if rollno is not None:
                              student = Student.objects.filter(rollno=rollno).first()

                        email = row[1].value
                        if email is not None:
                              email = email.strip()
                              if student is not None:
                                    user = student.user
                                    user.email = email
                                    user.save()
                              else:
                                    user = User.objects.filter(email=email).first()
                                    if user is None:
                                          user = User(email=email)
                                          password = ''.join(choice(string.ascii_letters + string.digits + '!@#$%^&*()[]{};:?.,<>_') for _ in range(8))
                                          user.set_password(password)
                                          user.save()
                                    try:
                                          if user.student is not None:
                                                student = user.student
                                    except:
                                          pass
                        
                        # current_room_type
                        current_room = row[8].value
                        is_current_room_null = False
                        if current_room is not None:
                              current_room = current_room.strip()
                              if current_room=='':
                                    # check if admin wants to make current room related field null
                                    is_current_room_null = True
                                    current_room = None
                        else:
                              is_current_room_null = True
                        
                        current_roomtype = None
                        current_hostel = row[7].value
                        if current_hostel is not None:
                              if is_current_room_null:
                                    current_hostel = None
                              else:
                                    current_hostel = current_hostel.strip()
                                    current_hostel = Hostel.objects.filter(name__iexact=current_hostel).first()
                                    if current_hostel is None:
                                          raise Exception(f'Current Hostel {current_hostel} not found!')
                                    current_roomtype = RoomType.objects.filter(Q(name__iexact=current_room) & Q(hostel = current_hostel)).first()
                                    if current_roomtype is None:
                                          raise Exception(f'Current Room Type {current_room} not found for Hostel {current_hostel}!')
                        
                        # alloted_room_type
                        alloted_room = row[10].value
                        is_alloted_room_null = False
                        if alloted_room is not None:
                              alloted_room = alloted_room.strip()
                              if alloted_room=='':
                                    # check if admin wants to make allocated room related details null
                                    is_alloted_room_null = True
                                    alloted_room = None
                        else:
                              is_alloted_room_null = True
                              
                        alloted_roomtype = None 
                        alloted_hostel = row[9].value
                        if alloted_hostel is not None:
                              if is_alloted_room_null:
                                    alloted_hostel = None
                              else:
                                    alloted_hostel = alloted_hostel.strip()
                                    alloted_hostel = Hostel.objects.filter(name__iexact=alloted_hostel).first()
                                    if alloted_hostel is None:
                                          raise Exception(f'Alloted Hostel {alloted_hostel} not found!')
                                    alloted_roomtype = RoomType.objects.filter(Q(name__iexact=alloted_room) & Q(hostel = alloted_hostel)).first()
                                    if alloted_roomtype is None:
                                          raise Exception(f'Alloted Room Type {alloted_room} not found!')

                        name = row[2].value
                        if (name is not None):
                              name = name.strip()

                        gender = row[6].value
                        if gender is not None:
                              gender = gender.strip()

                        if student is None:
                              # create new student
                              student = Student(
                                    name = name, 
                                    rollno = rollno, 
                                    phoneno = phoneno, 
                                    gender = gender, 
                                    cg = cg, 
                                    batch = batch, 
                                    user = user,
                                    current_room = current_roomtype,
                                    alloted_room = alloted_roomtype
                                    )
                              student.save()

                              group = Group(leader = student, cg = student.cg)
                              group.save()
                        else:
                              # update details of existing student
                              if name is not None:
                                    student.name = name
                              if rollno is not None:
                                    student.rollno = rollno
                              if phoneno is not None:
                                    student.phoneno = phoneno
                              if gender is not None:
                                    student.gender = gender
                              if cg is not None:
                                    student.cg = cg
                              if batch is not None:
                                    student.batch = batch
                              if current_roomtype is not None:
                                    student.current_room = current_roomtype
                              if alloted_roomtype is not None:
                                    student.alloted_room = alloted_roomtype
                              student.save()

                              group = Group.objects.filter(leader=student).first()
                              if group is None:
                                    group = Group(leader=student, cg=student.cg)
                                    group.save()
                              else:
                                    # update group cg
                                    updatedcg = student.cg
                                    groupSize = 1
                                    for member in group.members.all():
                                          updatedcg += member.cg
                                          groupSize += 1
                                    updatedcg /= groupSize
                                    group.cg = round(updatedcg, 2)
                                    group.save()

                        successCnt += 1
            
            except Exception as e:
                  # log error
                  errors.append(f'Creation of item {cnt} unsuccessful! Error: {str(e)}')
                  failureCnt += 1

            cnt += 1
      
      res = {
            'successful': successCnt,
            'unsuccessful': failureCnt,
            'errors': errors
      }

      end_time = datetime.now()
      time_diff = (end_time - start_time).total_seconds() / 60

      with open(log_file_path, 'a') as f:
            f.write(f'Execution completed in {time_diff} minutes.\n')
            f.write(json.dumps(res, indent=2))
            f.write('\n')
      
      errors_str = '\n- '.join(errors)
      
      subject = f'Execution results ready for the task Import Students'
      msg = f"""File: {filename}

Task Start Time: {start_time}
Task End Time: {end_time}
Time Elapsed: {time_diff}

Number of entries successfully executed: {successCnt}
Number entries failed: {failureCnt}

Errors
- {'No Errors' if len(errors)==0 else errors_str}
"""

      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()

      send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL), connection=connection, fail_silently=False)

      connection.close()

      return res


@app.task(name = "import_defaulters", queue='hms-admin-task')
def import_defaulters(filename):
      start_time = datetime.now()

      path = os.path.join(settings.BASE_DIR, 'imported-data', 'defaulter', filename)
      wb = load_workbook(path, data_only=True)
      ws = wb.active

      log_file_path = os.path.join(settings.LOGS_ROOT, 'import_defaulters.log')
      with open(log_file_path, 'a') as f:
            f.write(f"\nProcessing file {filename} at {start_time}...\n")

      # check if required fields are present in data
      if ws['A1'].value!='student':
            with open(log_file_path, 'a') as f:
                  f.write(f"Required 'student' as column name but found {ws['A1'].value} at cell A1.\n")
                  f.write("Aborting...\n\n")
            raise Exception(f"student field missing! Required 'student' as column value but found {ws['A1'].value} at cell A1.")
      
      cnt = 1
      successCnt = 0
      failureCnt = 0
      errors = []

      for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            try:
                  with transaction.atomic():
                        student = Student.objects.filter(rollno=row[0].value).first()
                        if student is None:
                              raise Exception(f"Student with roll number {row[0].value} not found!")
                        instance = Defaulter.objects.filter(student=student).first()
                        if instance is None:
                              instance = Defaulter(student=student)
                              instance.save()
                        successCnt += 1
            except BaseException as e:
                  errors.append(f"Creation of item {cnt} unsuccessful. Error: {str(e)}")
                  failureCnt += 1
            cnt += 1
      
      res = {
            'successful': successCnt,
            'unsuccessful': failureCnt,
            'errors': errors
      }

      end_time = datetime.now()
      time_diff = (end_time - start_time).total_seconds() / 60

      with open(log_file_path, 'a') as f:
            f.write(f'Execution completed in {time_diff} minutes.')
            f.write(json.dumps(res, indent=2))
            f.write('\n')
      
      errors_str = '\n- '.join(errors)
      
      subject = f'Execution results ready for the task Import Defaulters'
      msg = f"""File: {filename}

Task Start Time: {start_time}
Task End Time: {end_time}
Time Elapsed: {time_diff}

Number of entries successfully executed: {successCnt}
Number entries failed: {failureCnt}

Errors
- {'No Errors' if len(errors)==0 else errors_str}
"""

      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()

      send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL), connection=connection, fail_silently=False)

      connection.close()

      return res


@app.task(name = 'allot', queue='hms-admin-task')
def allot(section_ids):
      retained_students_cnt = 0
      retained_groups_cnt = 0

      alloted_students_cnt = 0
      alloted_groups_cnt = 0

      partial_allot_students_cnt = 0
      partial_allot_groups_cnt = 0

      logs_instance = AllotmentStatus()
      logs_instance.save()
      
      sections = []
      
      for section_id in section_ids:

            section = Section.objects.filter(id=section_id).first()
            sections.append(section)

            # storing the rooms capacity in a dictionary
            choices_queryset = section.choices.all()
            offered_room_types_capacity = {}
            for choice in choices_queryset:
                  offered_room_types_capacity[choice.room_type.id] = choice.capacity
            
            # Subtract Manually Alloted Students from capacity
            alloted_queryset = Student.objects.filter(batch=section.batch, gender=section.gender, alloted_room__isnull=False).select_related('leader_of_group', 'group').prefetch_related('leader_of_group__members', 'group__members').all()
            alloted_students_map = {}
            for student in alloted_queryset:
                  try:
                        gp = student.leader_of_group
                        if gp.members.count() > 0:
                              newleader = gp.members.first()
                              gp.leader = newleader

                              ms = gp.members.all()

                              updatedcg = gp.leader.cg  - student.cg
                              cnt = 0
                              for member in ms:
                                    updatedcg += member.cg
                                    cnt += 1
                              updatedcg /= cnt
                              gp.cg = round(updatedcg, 2)
                              gp.save()

                              newleader.group = None
                              newleader.save()

                              for member in ms:
                                    send_teamleader_change_mail.delay(gp.leader.name, gp.leader.rollno, member.user.email, member.name)
                              send_teamleader_change_mail.delay(gp.leader.name, gp.leader.rollno, gp.leader.user.email, gp.leader.name)

                              Group.objects.create(leader=student, cg=student.cg)
                  
                  except ObjectDoesNotExist:
                        if student.group is not None:
                              gp = student.group

                              left_group_mail.delay(gp.leader.name, student.name, student.rollno, gp.leader.user.email)
                              ms = gp.members.all()
                              for m in ms:
                                    if m != student:
                                          left_group_mail.delay(m.name, student.name, student.rollno, m.user.email)
                              
                              # due to prefetch, members were fetched before execution of above statement, so adjust cg of removed student
                              updatedcg = gp.leader.cg  - student.cg
                              cnt = 0
                              for member in ms:
                                    updatedcg += member.cg
                                    cnt += 1
                              updatedcg /= cnt
                              gp.cg = round(updatedcg, 2)
                              gp.save()

                              student.group = None
                              student.save()

                              Group.objects.create(leader=student, cg=student.cg)

                  offered_room_types_capacity[student.alloted_room.id] -= 1
                  alloted_students_map[student.id] = True
            
            # Handing groups that retained their previous room
            retain_groups = Group.objects.filter(is_retained=True, leader__batch=section.batch, leader__gender=section.gender).all()
            for group in retain_groups:
                  if offered_room_types_capacity[group.leader.current_room.id] > 0 and not alloted_students_map.get(group.leader.id, False):
                        group.leader.preview_room = group.leader.current_room
                        group.leader.save()
                        offered_room_types_capacity[group.leader.current_room.id] -= 1
                        retained_students_cnt += 1
                  for member in group.members.all():
                        if offered_room_types_capacity[member.current_room.id] > 0 and not alloted_students_map.get(member.id, False):
                              member.preview_room = member.current_room
                              member.save()
                              offered_room_types_capacity[member.current_room.id] -= 1
                              retained_students_cnt += 1
                  retained_groups_cnt += 1

            # Handling groups that have filled preferences
            unalloted_groups = []
            groups = Group.objects.filter(is_retained=False).annotate(
                  preferences_count = Count('preferences')
            ).filter(preferences_count__gt=0).order_by('-cg').all()
            for group in groups:
                  members = group.members.all()
                  group_size = len(members) + 1
                  preferences = Preference.objects.filter(group=group).order_by('priority').all()
                  got_allotment = False
                  for preference in preferences:
                        choice = preference.room_type_choice
                        cnt = 0
                        if offered_room_types_capacity[choice.room_type.id] < group_size:
                              continue
                        if not alloted_students_map.get(group.leader.id, False):
                              room_type = choice.room_type
                              group.leader.preview_room = room_type
                              group.leader.save()
                              cnt += 1
                        for member in group.members.all():
                              if not alloted_students_map.get(group.leader.id, False):
                                    member.preview_room = room_type
                                    member.save()
                                    cnt += 1
                        offered_room_types_capacity[choice.room_type.id] -= cnt
                        got_allotment = True
                        alloted_students_cnt += cnt
                        alloted_groups_cnt += 1
                        break
                  if not got_allotment:
                        unalloted_groups.append(group)
            
            # Handle groups that can't get same hostel
            for group in unalloted_groups:
                  leader = group.leader
                  students = [group.leader]
                  for member in group.members.all():
                        students.append(member)
                  students.sort(key = lambda x: x.cg, reverse = True)
                  preferences = group.preferences.order_by('priority').all()
                  ptr = 0
                  
                  for preference in preferences:
                        room_type = preference.room_type_choice.room_type
                        while ptr < len(students) and offered_room_types_capacity[room_type.id] > 0:
                              students[ptr].preview_room = room_type
                              students[ptr].save()
                              offered_room_types_capacity[room_type.id] -= 1
                              ptr += 1
                  
                  logs_leader = AllotmentLogsStudent.objects.create(
                        name = leader.name,
                        email = leader.user.email,
                        rollno = leader.rollno,
                        phoneno = leader.phoneno,
                        cg = leader.cg,
                        preview_room = leader.preview_room
                  )

                  logs_group = AllotmentLogsGroup.objects.create(
                        leader = logs_leader,
                        cg = group.cg,
                        status = logs_instance
                  )

                  for member in group.members.all():
                        AllotmentLogsStudent.objects.create(
                              name = leader.name,
                              email = member.user.email,
                              rollno = member.rollno,
                              phoneno = member.phoneno,
                              cg = member.cg,
                              preview_room = member.preview_room,
                              group = logs_group
                        )
                  
                  partial_allot_groups_cnt += 1
                  partial_allot_students_cnt += len(students)
            
            logs_instance.sections.add(section)

      logs_instance.retained_students_cnt = retained_students_cnt
      logs_instance.retained_groups_cnt = retained_groups_cnt
      logs_instance.alloted_students_cnt = alloted_students_cnt
      logs_instance.alloted_groups_cnt = alloted_groups_cnt
      logs_instance.partial_allot_students_cnt = partial_allot_students_cnt
      logs_instance.partial_allot_groups_cnt = partial_allot_groups_cnt
      logs_instance.save()

      from dashboard.serializers import AllotmentStatusSerializer
      
      serializer = AllotmentStatusSerializer(logs_instance)
      res = serializer.data

      subject = f'Execution results ready for the task Allotment'
      msg = f"""Successfully alloted hostels to students for section(s) {', '.join(f"{section.batch.name} {'Girls' if section.gender=='F' else 'Boys'}" for section in sections)}

Execution Results:
{json.dumps(res, indent=2)}"""

      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()

      send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL), connection=connection, fail_silently=False)

      connection.close()
      
      return res


@app.task(name = 'release_allotment_results', queue='hms-admin-task')
def release_allotment_results(section_id, fee_submission_deadline, reporting_information):
      section = Section.objects.filter(id=section_id).first()

      pending_allotment_students = Student.objects.filter(gender=section.gender, batch=section.batch, preview_room__isnull=False, alloted_room__isnull=True).all()
      for student in pending_allotment_students:
            student.alloted_room = student.preview_room
            student.save()
      
      retained_groups = Group.objects.filter(leader__gender=section.gender, leader__batch=section.batch, is_retained=True).all()
      for group in retained_groups:
            group.leader.alloted_room = group.leader.current_room
            group.leader.save()
            for member in group.members.all():
                  member.alloted_room = member.current_room
                  member.save()
      
      alloted_students = Student.objects.filter(gender=section.gender, batch=section.batch, alloted_room__isnull=False).all()
      
      academic_session = AcademicSession.objects.first()
      if academic_session is None:
            academic_session = ''
      else:
            academic_session = academic_session.name
      
      for group in retained_groups:
            students_list = [group.leader]
            for member in group.members.all():
                  students_list.append(member)
            
            for student in students_list:
                  subject = f"Hi {student.name}, Hostel Allotment Results are out for Academic Session {academic_session}"
      
                  idx = cache.get('emailIdIndex', 0)
                  cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
                  
                  connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
                  connection.open()
                  context = {
                        "name": student.name,
                        "fee": student.alloted_room.fee,
                        "academic_session": academic_session,
                        "fee_submission_deadline": fee_submission_deadline,
                        "reporting_information": reporting_information
                  }
                  html_message = render_to_string('dashboard/retain_result.html', context)
                  msg = strip_tags(html_message)
                  email = EmailMultiAlternatives(subject, msg, settings.EMAIL_HOST_USERS[idx], [student.user.email], reply_to = ['queries_studentaffairs@thapar.edu'], connection=connection)
                  email.attach_alternative(html_message, "text/html")
                  email.send()
                  
                  connection.close()
      
      for students_list in [alloted_students, pending_allotment_students]:
            for student in students_list:
                  subject = f"Hi {student.name}, Hostel Allotment Results are out for Academic Session {academic_session}"
      
                  idx = cache.get('emailIdIndex', 0)
                  cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
                  
                  connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
                  connection.open()
                  context = {
                        "name": student.name,
                        "hostel": student.alloted_room.hostel.name,
                        "room_type": student.alloted_room.name,
                        "academic_session": academic_session,
                        "fee": student.alloted_room.fee,
                        "fee_submission_deadline": fee_submission_deadline,
                        "reporting_information": reporting_information
                  }
                  html_message = render_to_string('dashboard/allotment_result.html', context)
                  msg = strip_tags(html_message)
                  email = EmailMultiAlternatives(subject, msg, settings.EMAIL_HOST_USERS[idx], [student.user.email], reply_to = ['queries_studentaffairs@thapar.edu'], connection=connection)
                  email.attach_alternative(html_message, "text/html")
                  email.send()
                  
                  connection.close()
      
      subject = f'Execution results ready for the task Release Allotment Results'
      msg = f"Successfully released Allotment Results for section {section.batch.name} {'Girls' if section.gender=='F' else 'Boys'} and sent mails regarding the same."

      idx = cache.get('emailIdIndex', 0)
      cache.set('emailIdIndex', (idx + 1) % settings.EMAIL_HOST_USERS_COUNT)
      
      connection = get_connection(username=settings.EMAIL_HOST_USERS[idx], password=settings.EMAIL_HOST_PASSWORDS[idx], fail_silently=False)
      connection.open()

      send_mail(subject, msg, settings.EMAIL_HOST_USERS[idx], (settings.ADMIN_EMAIL, settings.DEVELOPER_EMAIL), connection=connection, fail_silently=False)

      connection.close()

      return 1
